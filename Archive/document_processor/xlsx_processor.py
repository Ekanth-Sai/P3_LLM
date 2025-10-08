# document_processor/xlsx_processor.py

from openpyxl import load_workbook
import os
from typing import List
from PIL import Image

from .base_processor import BaseDocumentProcessor

class XlsxProcessor(BaseDocumentProcessor):
    """Processes .xlsx files."""

    def extract_text(self) -> str:
        workbook = load_workbook(self.file_path)
        full_text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            full_text.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                full_text.append("\t".join(row_values))
        return "\n".join(full_text)

    def extract_images(self) -> List[bytes]:
        # openpyxl does not directly support extracting embedded images easily
        # For a full solution, you might need to unzip the .xlsx file and parse XML
        # or use a more advanced library like 'pypandoc' or 'unoconv' if available
        print(f"Warning: Image extraction from .xlsx is not directly supported by openpyxl for {self.filename}.")
        return []